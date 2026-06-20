from pathlib import Path

from openpyxl import Workbook

from app.services.chunker import Chunker
from app.services.document_parser import DocumentParser
from app.services.table_parser import TableParser


def test_csv_parser_preserves_table_metadata_and_readable_text(tmp_path: Path) -> None:
    path = tmp_path / "products.csv"
    path.write_text("model,voltage,power\nA100,220V,500W\nA200,380V,1200W\n", encoding="utf-8")

    sections = DocumentParser().parse(path)

    assert len(sections) == 1
    section = sections[0]
    assert "File: products.csv" in section.text
    assert "Sheet: CSV" in section.text
    assert "Rows: 2-3" in section.text
    assert "Columns: model, voltage, power" in section.text
    assert "Row 2:" in section.text
    assert "model: A100" in section.text
    assert section.metadata == {
        "source_type": "table",
        "sheet_name": "CSV",
        "row_start": 2,
        "row_end": 3,
        "column_names": ["model", "voltage", "power"],
        "table_index": 0,
        "source_range": "2-3",
    }


def test_table_chunks_are_not_split_and_keep_metadata(tmp_path: Path) -> None:
    path = tmp_path / "products.csv"
    path.write_text("model,voltage\nA100,220V\n", encoding="utf-8")
    sections = DocumentParser().parse(path)

    chunks = Chunker(chunk_size=20, chunk_overlap=5).chunk(sections)

    assert len(chunks) == 1
    assert chunks[0].chunk_text.startswith("File: products.csv")
    assert chunks[0].metadata["source_type"] == "table"
    assert chunks[0].metadata["row_start"] == 2
    assert chunks[0].metadata["column_names"] == ["model", "voltage"]


def test_xlsx_parser_preserves_sheet_and_row_range(tmp_path: Path) -> None:
    path = tmp_path / "products.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ProductA"
    sheet.append(["model", "protocol"])
    sheet.append(["A100", "Modbus"])
    sheet.append(["A200", "EtherCAT"])
    workbook.save(path)

    sections = DocumentParser().parse(path)

    assert len(sections) == 1
    section = sections[0]
    assert "Sheet: ProductA" in section.text
    assert "Rows: 2-3" in section.text
    assert "protocol: EtherCAT" in section.text
    assert section.section_title == "ProductA"
    assert section.metadata["source_type"] == "table"
    assert section.metadata["sheet_name"] == "ProductA"
    assert section.metadata["row_start"] == 2
    assert section.metadata["row_end"] == 3


def test_xlsx_parser_handles_multiple_sheets_and_skips_empty_sheet(tmp_path: Path) -> None:
    path = tmp_path / "multi.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "Maintenance"
    first.append(["fault_code", "action"])
    first.append(["E12", "Check wiring"])
    second = workbook.create_sheet("Empty")
    third = workbook.create_sheet("Materials")
    third.append(["material_code", "supplier"])
    third.append(["MAT-001", "ACME"])
    workbook.save(path)

    sections = DocumentParser().parse(path)

    assert [section.metadata["sheet_name"] for section in sections] == ["Maintenance", "Materials"]
    assert "fault_code: E12" in sections[0].text
    assert "material_code: MAT-001" in sections[1].text
    assert second.title == "Empty"


def test_parser_normalizes_blank_and_duplicate_headers(tmp_path: Path) -> None:
    path = tmp_path / "headers.csv"
    path.write_text("model,model,\nA100,A100-alt,notes\n", encoding="utf-8")

    sections = DocumentParser().parse(path)

    assert sections[0].metadata["column_names"] == ["model", "model_2", "Column 3"]
    assert "model: A100" in sections[0].text
    assert "model_2: A100-alt" in sections[0].text
    assert "Column 3: notes" in sections[0].text


def test_parser_skips_empty_rows(tmp_path: Path) -> None:
    path = tmp_path / "empty_rows.csv"
    path.write_text("model,voltage\n,\nA200,380V\n", encoding="utf-8")

    sections = DocumentParser().parse(path)

    assert len(sections) == 1
    assert sections[0].metadata["row_start"] == 3
    assert sections[0].metadata["row_end"] == 3
    assert "Row 2:" not in sections[0].text
    assert "Row 3:" in sections[0].text


def test_parser_splits_long_tables_by_row_groups(tmp_path: Path) -> None:
    path = tmp_path / "long.csv"
    rows = ["model,description"]
    rows.extend(f"A{i}," + "x" * 80 for i in range(20))
    path.write_text("\n".join(rows), encoding="utf-8")

    parser = TableParser()
    parser.target_chunk_size = 250
    sections = parser.parse(path)

    assert len(sections) > 1
    assert all(section.metadata["source_type"] == "table" for section in sections)
    assert all("Columns: model, description" in section.text for section in sections)
    assert sections[0].metadata["row_start"] == 2
    assert sections[-1].metadata["row_end"] == 21
